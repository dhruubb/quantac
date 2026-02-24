"""
ingest_excel.py
Reads financial Excel files (Screener.in format) and converts them into
natural language text chunks for embedding into FAISS.

Handles: Profit & Loss, Quarterly, Balance Sheet, Cash Flow data
Works for: ICICI Bank, TCS, Infosys, Reliance Industries, Adani Power
"""

import openpyxl
from datetime import datetime


# ======================
# HELPERS
# ======================
def fmt_date(val):
    """Format a datetime to a readable year string."""
    if isinstance(val, datetime):
        month = val.month
        year = val.year
        # Indian fiscal year: Apr-Mar, so March 2025 = FY2024-25
        if month == 3:
            return f"FY{year-1}-{str(year)[2:]}"
        elif month in [6, 9, 12]:
            # Quarter ending
            quarter_map = {6: "Q1", 9: "Q2", 12: "Q3"}
            fy_year = year if month != 12 else year
            return f"{quarter_map[month]} FY{fy_year-1}-{str(fy_year)[2:]}" if month != 12 \
                else f"Q3 FY{year}-{str(year+1)[2:]}"
        return str(year)
    return str(val) if val else ""


def fmt_num(val, unit="crore"):
    """Format a number with Indian financial conventions."""
    if val is None:
        return "N/A"
    try:
        val = float(val)
        if abs(val) >= 100000:
            return f"₹{val/100:.0f} crore"  # already in crore, just format
        return f"₹{val:,.2f} crore"
    except (TypeError, ValueError):
        return str(val)


def pct_change(new, old):
    """Calculate percentage change between two values."""
    try:
        if old and old != 0:
            change = ((float(new) - float(old)) / abs(float(old))) * 100
            direction = "up" if change >= 0 else "down"
            return f"{direction} {abs(change):.1f}%"
    except (TypeError, ValueError):
        pass
    return ""


# ======================
# PARSE DATA SHEET
# ======================
def parse_data_sheet(ws):
    """
    Parse the raw Data Sheet into structured sections.
    Returns dict with sections: P&L, Quarters, Balance Sheet, Cash Flow, Meta
    """
    data = {
        "meta": {},
        "pl": {"dates": [], "rows": {}},
        "quarters": {"dates": [], "rows": {}},
        "bs": {"dates": [], "rows": {}},
        "cf": {"dates": [], "rows": {}},
    }

    current_section = None
    section_map = {
        "PROFIT & LOSS": "pl",
        "QUARTERS": "quarters",
        "BALANCE SHEET": "bs",
        "CASH FLOW:": "cf",
    }

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue

        label = str(row[0]).strip() if row[0] else ""
        values = list(row[1:])

        # Detect section headers
        if label.upper() in section_map:
            current_section = section_map[label.upper()]
            continue

        # Meta data
        if label == "COMPANY NAME":
            data["meta"]["company"] = values[0] if values else ""
            continue
        if label == "Current Price":
            data["meta"]["current_price"] = values[0]
            continue
        if label == "Market Capitalization":
            data["meta"]["market_cap"] = values[0]
            continue

        # Date rows
        if label == "Report Date" and current_section:
            dates = [fmt_date(v) for v in values if v is not None]
            data[current_section]["dates"] = dates
            continue

        # Data rows
        if current_section and label and label not in ["META", "DERIVED:", "PRICE:"]:
            # Filter out None values but keep alignment with dates
            data[current_section]["rows"][label] = values

    return data


# ======================
# GENERATE P&L CHUNKS
# ======================
def generate_pl_chunks(data, company, recent_years=5):
    """Generate natural language chunks from P&L data."""
    chunks = []
    pl = data["pl"]
    dates = pl["dates"]
    rows = pl["rows"]

    if not dates:
        return chunks

    # Focus on recent years
    recent_dates = dates[-recent_years:]
    recent_idx = len(dates) - recent_years

    # Key metrics to narrate
    key_metrics = ["Sales", "Net profit", "Profit before tax", "Interest",
                   "Employee Cost", "Depreciation", "Tax"]

    # --- Annual summary chunk ---
    lines = [f"{company} — Annual Profit & Loss Summary (₹ Crore)\n"]
    lines.append(f"Years covered: {', '.join(recent_dates)}\n")

    for metric in key_metrics:
        if metric not in rows:
            continue
        vals = rows[metric][recent_idx:]
        vals = [v for v in vals if v is not None]
        if not vals:
            continue

        row_dates = recent_dates[-len(vals):]
        val_strs = [f"{d}: {fmt_num(v)}" for d, v in zip(row_dates, vals)]
        lines.append(f"{metric}: {' | '.join(val_strs)}")

        # Add YoY change for most recent year
        if len(vals) >= 2:
            chg = pct_change(vals[-1], vals[-2])
            if chg:
                lines.append(f"  → {metric} was {chg} YoY in {row_dates[-1]}")

    chunks.append({
        "content": "\n".join(lines),
        "metadata": {
            "company": company,
            "year": "Multi-Year",
            "doc_type": "Financial Data",
            "section": "Profit & Loss",
            "chunk_index": 0
        }
    })

    # --- Per-year detailed chunks ---
    for i, date in enumerate(recent_dates):
        actual_idx = recent_idx + i
        lines = [f"{company} — Profit & Loss for {date} (₹ Crore)\n"]

        for metric in key_metrics:
            if metric not in rows:
                continue
            vals = rows[metric]
            if actual_idx >= len(vals) or vals[actual_idx] is None:
                continue
            val = vals[actual_idx]
            prev_val = vals[actual_idx - 1] if actual_idx > 0 else None
            chg = f" ({pct_change(val, prev_val)} YoY)" if prev_val else ""
            lines.append(f"• {metric}: {fmt_num(val)}{chg}")

        # Sales growth narrative
        if "Sales" in rows and "Net profit" in rows:
            sales = rows["Sales"]
            net = rows["Net profit"]
            if actual_idx < len(sales) and sales[actual_idx]:
                margin = (float(net[actual_idx]) / float(sales[actual_idx])) * 100 \
                    if net[actual_idx] else 0
                lines.append(f"• Net Profit Margin: {margin:.1f}%")

        chunks.append({
            "content": "\n".join(lines),
            "metadata": {
                "company": company,
                "year": date,
                "doc_type": "Financial Data",
                "section": "Profit & Loss",
                "chunk_index": i + 1
            }
        })

    return chunks


# ======================
# GENERATE QUARTERLY CHUNKS
# ======================
def generate_quarterly_chunks(data, company):
    """Generate natural language chunks from quarterly data."""
    chunks = []
    q = data["quarters"]
    dates = q["dates"]
    rows = q["rows"]

    if not dates:
        return chunks

    key_metrics = ["Sales", "Net profit", "Profit before tax",
                   "Operating Profit", "Interest", "Expenses"]

    # Recent 8 quarters
    recent_dates = dates[-8:]
    recent_idx = max(0, len(dates) - 8)

    lines = [f"{company} — Quarterly Financial Data (₹ Crore)\n"]
    lines.append(f"Quarters: {', '.join(recent_dates)}\n")

    for metric in key_metrics:
        if metric not in rows:
            continue
        vals = rows[metric][recent_idx:]
        vals_clean = [v for v in vals if v is not None]
        if not vals_clean:
            continue
        row_dates = recent_dates[-len(vals_clean):]
        val_strs = [f"{d}: {fmt_num(v)}" for d, v in zip(row_dates, vals_clean)]
        lines.append(f"{metric}: {' | '.join(val_strs)}")

    # Latest quarter narrative
    if dates:
        last_q = recent_dates[-1]
        lines.append(f"\nLatest Quarter ({last_q}) highlights:")
        for metric in ["Sales", "Net profit", "Operating Profit"]:
            if metric in rows:
                vals = rows[metric]
                if vals and vals[-1] is not None:
                    prev = vals[-5] if len(vals) >= 5 else None  # vs same quarter last year
                    chg = f" ({pct_change(vals[-1], prev)} vs same quarter last year)" if prev else ""
                    lines.append(f"• {metric}: {fmt_num(vals[-1])}{chg}")

    chunks.append({
        "content": "\n".join(lines),
        "metadata": {
            "company": company,
            "year": "Quarterly",
            "doc_type": "Financial Data",
            "section": "Quarterly Results",
            "chunk_index": 0
        }
    })

    return chunks


# ======================
# GENERATE BALANCE SHEET CHUNKS
# ======================
def generate_bs_chunks(data, company, recent_years=5):
    """Generate natural language chunks from balance sheet data."""
    chunks = []
    bs = data["bs"]
    dates = bs["dates"]
    rows = bs["rows"]

    if not dates:
        return chunks

    recent_dates = dates[-recent_years:]
    recent_idx = len(dates) - recent_years

    key_metrics = ["Equity Share Capital", "Reserves", "Borrowings",
                   "Total", "Investments", "Cash & Bank", "Other Assets",
                   "Return on Equity", "Return on Capital Emp"]

    lines = [f"{company} — Balance Sheet Summary (₹ Crore)\n"]
    lines.append(f"Years: {', '.join(recent_dates)}\n")

    for metric in key_metrics:
        if metric not in rows:
            continue
        vals = rows[metric][recent_idx:]
        vals_clean = [v for v in vals if v is not None]
        if not vals_clean:
            continue
        row_dates = recent_dates[-len(vals_clean):]

        # Format ratios differently
        if metric in ["Return on Equity", "Return on Capital Emp"]:
            val_strs = [f"{d}: {float(v):.1f}%" if v else f"{d}: N/A"
                        for d, v in zip(row_dates, vals_clean)]
        else:
            val_strs = [f"{d}: {fmt_num(v)}" for d, v in zip(row_dates, vals_clean)]

        lines.append(f"{metric}: {' | '.join(val_strs)}")

        if len(vals_clean) >= 2:
            chg = pct_change(vals_clean[-1], vals_clean[-2])
            if chg:
                lines.append(f"  → {metric} {chg} YoY in {row_dates[-1]}")

    chunks.append({
        "content": "\n".join(lines),
        "metadata": {
            "company": company,
            "year": "Multi-Year",
            "doc_type": "Financial Data",
            "section": "Balance Sheet",
            "chunk_index": 0
        }
    })

    return chunks


# ======================
# GENERATE CASH FLOW CHUNKS
# ======================
def generate_cf_chunks(data, company, recent_years=5):
    """Generate natural language chunks from cash flow data."""
    chunks = []
    cf = data["cf"]
    dates = cf["dates"]
    rows = cf["rows"]

    if not dates:
        return chunks

    recent_dates = dates[-recent_years:]
    recent_idx = len(dates) - recent_years

    key_metrics = ["Cash from Operating Activity", "Cash from Investing Activity",
                   "Cash from Financing Activity", "Net Cash Flow"]

    lines = [f"{company} — Cash Flow Summary (₹ Crore)\n"]
    lines.append(f"Years: {', '.join(recent_dates)}\n")

    for metric in key_metrics:
        if metric not in rows:
            continue
        vals = rows[metric][recent_idx:]
        vals_clean = [v for v in vals if v is not None]
        if not vals_clean:
            continue
        row_dates = recent_dates[-len(vals_clean):]
        val_strs = [f"{d}: {fmt_num(v)}" for d, v in zip(row_dates, vals_clean)]
        lines.append(f"{metric}: {' | '.join(val_strs)}")

    chunks.append({
        "content": "\n".join(lines),
        "metadata": {
            "company": company,
            "year": "Multi-Year",
            "doc_type": "Financial Data",
            "section": "Cash Flow",
            "chunk_index": 0
        }
    })

    return chunks


# ======================
# MAIN ENTRY POINT
# ======================
def ingest_excel(excel_path, company):
    """
    Read an Excel file and return a list of text chunks with metadata.

    Args:
        excel_path : Path to the .xlsx file
        company    : Company name string (e.g. "ICICI Bank")

    Returns:
        List of dicts with 'content' and 'metadata' keys
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"   ❌ Failed to open {excel_path}: {e}")
        return []

    if "Data Sheet" not in wb.sheetnames:
        print(f"   ⚠️  No 'Data Sheet' found in {excel_path}")
        return []

    ws = wb["Data Sheet"]
    data = parse_data_sheet(ws)

    chunks = []
    chunks.extend(generate_pl_chunks(data, company))
    chunks.extend(generate_quarterly_chunks(data, company))
    chunks.extend(generate_bs_chunks(data, company))
    chunks.extend(generate_cf_chunks(data, company))

    return chunks


# ======================
# STANDALONE TEST
# ======================
if __name__ == "__main__":
    import sys

    excel_path = sys.argv[1] if len(sys.argv) > 1 else "icici/ICICI Bank.xlsx"
    company = sys.argv[2] if len(sys.argv) > 2 else "ICICI Bank"

    print(f"📊 Reading: {excel_path}")
    chunks = ingest_excel(excel_path, company)

    print(f"\n✅ Total chunks generated: {len(chunks)}")
    for c in chunks:
        print(f"\n--- {c['metadata']['section']} | {c['metadata']['year']} ---")
        print(c["content"][:600])
        print("...")